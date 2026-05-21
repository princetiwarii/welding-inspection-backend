"""
Report Service — PDF and Excel report generation
Generates inspection reports and uploads to S3.
"""
import io
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from fastapi import HTTPException

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import mm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.models.models import Inspection, AIResult, Defect, Report, ReportFormat
from app.utils.s3 import upload_file_to_s3
from app.utils.id_generator import generate_report_filename
from app.core.config import settings


class ReportService:

    @staticmethod
    async def generate_report(
        db: AsyncSession, inspection_id: str, fmt: str, user_id: int
    ) -> Report:
        # Load inspection with related data
        insp_result = await db.execute(
            select(Inspection).where(Inspection.inspection_id == inspection_id)
        )
        inspection = insp_result.scalar_one_or_none()
        if not inspection:
            raise HTTPException(status_code=404, detail="Inspection not found")

        ai_result_q = await db.execute(
            select(AIResult).where(AIResult.inspection_id == inspection_id)
        )
        ai_result = ai_result_q.scalar_one_or_none()

        defects = []
        if ai_result:
            defects_q = await db.execute(
                select(Defect).where(Defect.ai_result_id == ai_result.id)
            )
            defects = defects_q.scalars().all()

        fmt_upper = fmt.upper()
        if fmt_upper == "PDF":
            file_bytes, mime_type = ReportService._generate_pdf(inspection, ai_result, defects)
        elif fmt_upper == "EXCEL":
            file_bytes, mime_type = ReportService._generate_excel(inspection, ai_result, defects)
        else:
            raise HTTPException(status_code=400, detail="Unsupported format. Use PDF or EXCEL")

        file_name = generate_report_filename(inspection_id, fmt_upper)
        s3_key = f"{settings.REPORT_STORAGE_PREFIX}{inspection_id}/{file_name}"
        s3_url = upload_file_to_s3(file_bytes, s3_key, mime_type)

        report = Report(
            inspection_id=inspection_id,
            report_format=ReportFormat.PDF if fmt_upper == "PDF" else ReportFormat.EXCEL,
            s3_key=s3_key,
            s3_url=s3_url,
            file_name=file_name,
            file_size_bytes=len(file_bytes),
            generated_by=user_id,
        )
        db.add(report)
        await db.flush()
        return report

    @staticmethod
    def _generate_pdf(inspection, ai_result, defects) -> tuple:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
        elements.append(Paragraph("Welding Inspection Report", title_style))
        elements.append(Paragraph(f"Inspection ID: {inspection.inspection_id}", styles["Normal"]))
        elements.append(Spacer(1, 10*mm))

        # Inspection Info Table
        info_data = [
            ["Field", "Value"],
            ["Inspection ID", inspection.inspection_id],
            ["Object ID", inspection.object_id],
            ["Welding Position", str(inspection.welding_position or "N/A")],
            ["Status", str(inspection.status)],
            ["Overall Result", str(inspection.overall_result)],
            ["Scan Length (m)", str(inspection.scan_length_meters or "N/A")],
            ["Remarks", str(inspection.remarks or "None")],
            ["Submitted At", str(inspection.submitted_at or "N/A")],
            ["Completed At", str(inspection.completed_at or "N/A")],
        ]
        info_table = Table(info_data, colWidths=[60*mm, 110*mm])
        info_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10*mm))

        # AI Results Section
        if ai_result:
            elements.append(Paragraph("AI Defect Analysis", styles["Heading2"]))
            elements.append(Spacer(1, 3*mm))

            if defects:
                defect_data = [["#", "Type", "Severity", "Length(mm)", "Depth(mm)", "Accuracy%", "Description"]]
                for i, d in enumerate(defects, 1):
                    defect_data.append([
                        str(i),
                        d.defect_type,
                        str(d.severity),
                        str(d.length_mm or "-"),
                        str(d.depth_mm or "-"),
                        f"{d.ai_accuracy_pct}%" if d.ai_accuracy_pct else "-",
                        (d.description or "")[:60],
                    ])
                defect_table = Table(defect_data, colWidths=[10*mm, 28*mm, 22*mm, 22*mm, 22*mm, 20*mm, 46*mm])
                severity_colors = {"Low": "#d4edda", "Medium": "#fff3cd", "High": "#f8d7da"}
                style = TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d4a7a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("PADDING", (0, 0), (-1, -1), 3),
                ])
                defect_table.setStyle(style)
                elements.append(defect_table)
            else:
                elements.append(Paragraph("No defects detected.", styles["Normal"]))
        else:
            elements.append(Paragraph("AI analysis not yet completed.", styles["Normal"]))

        elements.append(Spacer(1, 8*mm))
        elements.append(Paragraph(f"Report generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))

        doc.build(elements)
        return buffer.getvalue(), "application/pdf"

    @staticmethod
    def _generate_excel(inspection, ai_result, defects) -> tuple:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspection Report"

        header_fill = PatternFill("solid", fgColor="1A365D")
        header_font = Font(color="FFFFFF", bold=True)

        # Inspection Info
        ws.append(["WELDING INSPECTION REPORT"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Field", "Value"])
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill

        info_rows = [
            ("Inspection ID", inspection.inspection_id),
            ("Object ID", inspection.object_id),
            ("Welding Position", str(inspection.welding_position or "N/A")),
            ("Status", str(inspection.status)),
            ("Overall Result", str(inspection.overall_result)),
            ("Scan Length (m)", str(inspection.scan_length_meters or "N/A")),
            ("Remarks", str(inspection.remarks or "None")),
        ]
        for row in info_rows:
            ws.append(row)

        ws.append([])
        ws.append(["AI DEFECT ANALYSIS"])
        ws.append(["#", "Defect Type", "Severity", "Length(mm)", "Depth(mm)", "Count", "Accuracy%", "Position", "Description"])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill

        severity_fills = {
            "Low": PatternFill("solid", fgColor="D4EDDA"),
            "Medium": PatternFill("solid", fgColor="FFF3CD"),
            "High": PatternFill("solid", fgColor="F8D7DA"),
        }
        for i, d in enumerate(defects, 1):
            row_idx = ws.max_row + 1
            ws.append([
                i, d.defect_type, str(d.severity),
                d.length_mm, d.depth_mm, d.count,
                d.ai_accuracy_pct, d.position, d.description
            ])
            sev = str(d.severity).split(".")[-1] if "." in str(d.severity) else str(d.severity)
            fill = severity_fills.get(sev)
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
